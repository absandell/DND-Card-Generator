"""XLSX reading and card generation orchestration."""
import math
import sys
from pathlib import Path

import openpyxl

from .config import CARDS_PER_PAGE
from .builder import _page_div, combined_html
from .renderers import render_spell_card, render_weapon_card, render_feature_card, render_item_card

RENDERERS = {
    "spell":   render_spell_card,
    "weapon":  render_weapon_card,
    "feature": render_feature_card,
    "item":    render_item_card,
}


def detect_sheet_type(name: str) -> str | None:
    n = name.lower()
    if "spell"                     in n: return "spell"
    if "weapon"                    in n: return "weapon"
    if "feature" in n or "trait"   in n: return "feature"
    if "item"                      in n: return "item"
    return None


def read_rows(ws) -> list[dict]:
    headers = [str(c.value or "").strip() for c in ws[1]]
    rows = []
    for raw in ws.iter_rows(min_row=2, values_only=True):
        record = {h: (str(v).strip() if v is not None else "") for h, v in zip(headers, raw)}
        if any(record.values()):
            rows.append(record)
    return rows


def generate(xlsx_path: Path, output_dir: Path) -> None:
    """Read the workbook and write one HTML file per card type to output_dir."""
    if not xlsx_path.exists():
        sys.exit(f"Error: '{xlsx_path}' not found.")

    output_dir.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.load_workbook(xlsx_path)

    total_cards = 0

    for sheet_name in wb.sheetnames:
        stype = detect_sheet_type(sheet_name)
        if stype is None:
            print(f"  [-] '{sheet_name}' -- unrecognised sheet name, skipping")
            continue

        rows = read_rows(wb[sheet_name])
        if not rows:
            print(f"  [-] '{sheet_name}' -- empty, skipping")
            continue

        renderer  = RENDERERS[stype]
        card_html = [renderer(r) for r in rows]
        num_pages = math.ceil(len(card_html) / CARDS_PER_PAGE)

        page_divs = [
            _page_div(card_html[p * CARDS_PER_PAGE : (p + 1) * CARDS_PER_PAGE])
            for p in range(num_pages)
        ]

        slug     = stype  # spell / weapon / feature / item
        out_path = output_dir / f"{slug}s.html"
        out_path.write_text(combined_html(page_divs), encoding="utf-8")

        print(f"  [{sheet_name}]  {len(rows)} cards, {num_pages} pages  ->  {out_path.name}")
        total_cards += len(rows)

    print(f"\n[+] {total_cards} cards written to '{output_dir}/'")
    print("    Open any file in a browser -> Ctrl+P -> Save as PDF")
